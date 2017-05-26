from openpyxl import utils
from report import evolution


class Entity:
    def __init__(self, make: str, model: str, version: str, prod_yr: int, model_yr: int, msrp: float, sample_date: int,
                 inc_code: str, jato_value: float, take_rate: float, manuf_contrib: float, interest_perc: float,
                 deposit_perc: float, max_term: int, internal_comments: str, public_notes: str,
                 volume: int, value: float):
        self.make = make
        self.model = model
        self.version = version
        self.prod_yr = prod_yr
        self.model_yr = model_yr
        self.msrp = msrp
        self.sample_date = sample_date
        self.inc_code = inc_code
        self.jato_value = jato_value
        self.take_rate = take_rate
        self.manuf_contrib = manuf_contrib
        self.interest_perc = interest_perc
        self.deposit_perc = deposit_perc
        self.max_term = max_term
        self.internal_comments = internal_comments
        self.public_notes = public_notes
        self.volume = volume
        self.value = value


class EquipmentAndIncentives(evolution.EvolutionReport):
    def __init__(self, data: []):
        def create_header(name: str, number_format: str, offset,
                          make_summary_formula: str, make_summary_formula_ranges: (tuple, ),
                          model_summary_formula: str, model_summary_formula_ranges: (tuple,)):
            header = self.Header(header_name=name, number_format=number_format, offset=offset)
            header.make_header = header.Summary(make_summary_formula, make_summary_formula_ranges)
            header.model_header = header.Summary(model_summary_formula, model_summary_formula_ranges)
            return header

        super().__init__()
        self.entities = [Entity(*result) for result in data]
        self.sample_headers.append(create_header(name='MSRP', number_format='#,###', offset=0,
                                                 make_summary_formula='', make_summary_formula_ranges=[(0, 0), ],
                                                 model_summary_formula='=IF(ISERROR(AVERAGE({0})),"-",AVERAGE({0}))',
                                                 model_summary_formula_ranges=[(1, 0)]))
        self.sample_headers.append(create_header(name='Equip. Value', number_format='#,###', offset=1,
                                                 make_summary_formula='', make_summary_formula_ranges=[(0, 0), ],
                                                 model_summary_formula='=IF(ISERROR(AVERAGE({0})),"-",AVERAGE({0}))',
                                                 model_summary_formula_ranges=[(1, 0)]))
        self.sample_headers.append(create_header(name='Manuf. Contrib.', number_format='#,###', offset=2,
                                                 make_summary_formula='', make_summary_formula_ranges=[(0, 0), ],
                                                 model_summary_formula='=IF(ISERROR(AVERAGE({0})),"-",AVERAGE({0}))',
                                                 model_summary_formula_ranges=[(1, 0)]))
        self.sample_headers.append(create_header(name='Finance Incentive', number_format='#,###', offset=3,
                                                 make_summary_formula='', make_summary_formula_ranges=[(0, 0), ],
                                                 model_summary_formula='=IF(ISERROR(AVERAGE({0})),"-",AVERAGE({0}))',
                                                 model_summary_formula_ranges=[(1, 0)]))
        self.sample_headers.append(create_header(name='Net Price', number_format='#,###', offset=4,
                                                 make_summary_formula='', make_summary_formula_ranges=[(0, 0), ],
                                                 model_summary_formula='=IF(ISERROR(AVERAGE({0})),"-",AVERAGE({0}))',
                                                 model_summary_formula_ranges=[(1, 0)]))
        self.sample_headers.append(create_header(name='Volume', number_format='#,###', offset=5,
                                                 make_summary_formula='', make_summary_formula_ranges=[(0, 0), ],
                                                 model_summary_formula='=IF(ISERROR(AVERAGE({0})),"-",AVERAGE({0}))',
                                                 model_summary_formula_ranges=[(1, 0)]))

    def __str__(self):
        from datetime import datetime
        return '{}.{}_{}'.format(str(super()), 'Equipment_and_Incentives_Report', datetime.today().strftime('%Y%m%d'))

    def generate_report(self):
        temp_ent = Entity(make='', model='', version='', prod_yr=0, model_yr=0, msrp=0, sample_date=0, inc_code='',
                          jato_value=0, take_rate=0, manuf_contrib=0, interest_perc=0, deposit_perc=0, max_term=0,
                          internal_comments='', public_notes='', volume=0, value=0)
        row_index = 0
        self.matrix = [[0 for l in range(200)] for c in range(31)]

        def on_make_change(new_make_name: str):
            if not temp_ent.make:
                self.fill_empty_vehicle_cells(version_row=row_index)
            temp_ent.make = new_make_name
            temp_ent.model = None
            self.write_make_header(make_name=new_make_name,
                                   amount_of_distinct_vehicles_of_make=0,
                                   amount_of_distinct_models_of_make=0,
                                   make_header_row=row_index)

        def on_model_change(new_model_name: str):
            if not temp_ent.model:
                self.fill_empty_vehicle_cells(version_row=row_index)
            temp_ent.model = new_model_name
            temp_ent.version = None
            self.write_model_header(model_name=new_model_name,
                                    amount_of_distinct_vehicles_of_model=0,
                                    model_header_row=row_index)

        def on_version_change(new_version_name: str):
            if not temp_ent.version:
                self.fill_empty_vehicle_cells(version_row=row_index)
            temp_ent.version = new_version_name
            temp_ent.prod_yr = None
            self.matrix[row_index, self.vehicle_desc_mark_up_col] = 'v'
            self.matrix[row_index, self.POSITION['vehicle_col']] = new_version_name

        def on_model_year_change(new_prod_yr: int, new_model_yr: int):
            if not temp_ent.prod_yr:
                self.fill_empty_vehicle_cells(version_row=row_index)
            temp_ent.prod_yr = new_prod_yr
            self.matrix[row_index, self.POSITION['prod_model_year_col']] = '{}/{}'.format(str(new_prod_yr % 100),
                                                                                          str(new_model_yr % 100))

        def write_vehicle_data(vehicle_ent: Entity, version_row: int):
            sample_date_index = self.sample_headers.index(vehicle_ent.sample_date)
            column_index = self.POSITION['first_sample_col'] + sample_date_index * len(self.sample_headers)
            self.matrix[version_row, column_index] = vehicle_ent.msrp
            self.matrix[version_row, column_index + 1] = vehicle_ent.value if vehicle_ent.value else 0
            self.matrix[version_row, column_index + 2] = vehicle_ent.manuf_contrib
            self.matrix[version_row, column_index + 3] = vehicle_ent.jato_value * vehicle_ent.take_rate
            self.matrix[version_row, column_index + 4] = '={}-{}-{}-{}'.format(
                '{}{}'.format(utils.get_column_letter(column_index), str(version_row)),
                '{}{}'.format(utils.get_column_letter(column_index + 1), str(version_row)),
                '{}{}'.format(utils.get_column_letter(column_index + 2), str(version_row)),
                '{}{}'.format(utils.get_column_letter(column_index + 3), str(version_row)))
            self.matrix[version_row, column_index + 5] = vehicle_ent.volume

        if not self.entities:
            raise IndexError('Cant generate report without data')

        self.sample_dates = sorted(set(entity.sample_date for entity in self.entities))
        self.vehicle_desc_mark_up_col = self.POSITION['first_sample_col'] + \
                                        len(self.sample_dates) * len(self.sample_headers) + 1
        for entity in self.entities:
            if temp_ent.make != entity.make:
                on_make_change(new_make_name=entity.make)
            if temp_ent.model != entity.model:
                on_model_change(new_model_name=entity.model)
                row_index += 1
            if temp_ent.version != entity.version:
                row_index += 1
                on_version_change(new_version_name=entity.version)
            if temp_ent.prod_yr != entity.prod_yr:
                row_index += 1
                on_model_year_change(new_prod_yr=entity.prod_yr, new_model_yr=entity.model_yr)
            write_vehicle_data(vehicle_ent=entity, version_row=row_index)

        self.finish_worksheet(last_row_index=row_index)
