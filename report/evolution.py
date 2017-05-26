import abc
from openpyxl import styles
from openpyxl import utils
from report import base


class EvolutionReport(base.Report, abc.ABC):
    def __init__(self):
        self.sample_dates = [int]
        self.sample_headers = [self.Header]
        self.vehicle_desc_mark_up_col = -1
        super().__init__()

    def __str__(self):
        return 'evolution'

    POSITION = {
        'title_header_row': 0,
        'time_header_row': 1,
        'info_header_row': 2,
        'first_sample_row': 3,
        'vehicle_col': 0,
        'prod_model_year_col': 1,
        'first_sample_col': 2,
        'summary_make_col': 0,
        'summary_time_header_row': 0
    }

    class Header(base.Report.Header):
        def __init__(self, header_name: str, number_format: str, offset: int):
            super().__init__(header_name=header_name, number_format=number_format, offset=offset)
            self.make_summary = self.Summary('', None)
            self.model_summary = self.Summary('', None)

        class Summary:
            def __init__(self, formula: str, formula_ranges: (tuple, )):
                """
                The formula to bring up a value in the summary
                :param formula: the formula to bring a value
                :param formula_ranges: a tuple with the up-most cells of vehicles to fall in the formula range relative 
                  to the summary cell
                """
                self.formula = formula
                self.formula_ranges = formula_ranges

            # TODO test
            def mount(self, cur_col: int, cur_row: int, distinct_sum_of_vehicles_of_model: int):
                if '{0}' not in self.formula:
                    return None
                return self.formula.format(*tuple([
                    '{}:{}'.format(  # formula cells range
                        '{}{}'.format(utils.get_column_letter(cur_col + formula_range[1]), cur_row + formula_range[0]),
                        '{}{}'.format(utils.get_column_letter(cur_col + formula_range[1]), cur_row + formula_range[0] +
                                      distinct_sum_of_vehicles_of_model - 1))
                    for formula_range in self.formula_ranges]))

    def fill_empty_vehicle_cells(self, version_row: int):
        """
        Fill the empty cells of the specified row so the analyst can see that those cells were looked for 
        :param version_row: 
        :return: 
        """
        for sample_date_index in range(0, len(self.sample_dates)):
            if not self.matrix[version_row][self.POSITION['first_sample_col']
                    + len(self.sample_headers) * sample_date_index]:
                for header in self.sample_headers:
                    self.matrix[version_row][self.POSITION['first_sample_col']
                                             + len(self.sample_headers) * sample_date_index + header.offset] = '?'

    def write_make_header(self, make_name: str, amount_of_distinct_vehicles_of_make: int,
                          amount_of_distinct_models_of_make: int, make_header_row: int):
        self.matrix[make_header_row, self.POSITION['vehicle_col']] = make_name
        for sample_date_index in range(len(self.sample_dates)):
            first_column_of_sample = self.POSITION['first_sample_col'] + sample_date_index * len(self.sample_headers)
            for header in self.sample_headers:
                absolute_column_of_header = first_column_of_sample + header.offset
                self.matrix[make_header_row, absolute_column_of_header] = header.make_summary.mount(
                    absolute_column_of_header, make_header_row,
                    amount_of_distinct_vehicles_of_make + amount_of_distinct_models_of_make)

    def write_model_header(self, model_name: str, amount_of_distinct_vehicles_of_model: int, model_header_row: int):
        self.matrix[model_header_row, self.POSITION['vehicle_col']] = model_name
        for sample_date_index in range(len(self.sample_dates)):
            first_column_of_sample = self.POSITION['first_sample_col'] + sample_date_index * len(self.sample_headers)
            for header in self.sample_headers:
                absolute_column_of_header = first_column_of_sample + header.offset
                self.matrix[model_header_row, absolute_column_of_header] = header.model_summary.mount(
                    absolute_column_of_header, model_header_row, amount_of_distinct_vehicles_of_model)

    def finish_worksheet(self, last_row_index: int):
        def add_header_xl_features():
            """
            Add color, font and size to the report headers
            :return: 
            """
            def all_headers_font(font: styles.Font):
                """
                Common features shared by all headers (font)
                :param font: 
                :return: 
                """
                self.ws['{}:{}'.format(
                    '{}{}'.format(utils.get_column_letter(self.POSITION['vehicle_col']),
                                  self.POSITION['title_header_row']),
                    '{}{}'.format(utils.get_column_letter(self.vehicle_desc_mark_up_col),
                                  self.POSITION['info_header_row']))].font = font

            def time_info_headers(border: styles.Border, interior_color: styles.Color, font: styles.Font):
                """
                Add features to the info headers
                :param border: 
                :param interior_color: 
                :param font: 
                :return: 
                """
                rang = self.ws['{}:{}'.format(
                    '{}{}'.format(utils.get_column_letter(self.POSITION['vehicle_col']),
                                  self.POSITION['time_header_row']),
                    '{}{}'.format(utils.get_column_letter(self.vehicle_desc_mark_up_col),
                                  self.POSITION['info_header_row']))]
                rang.border = border
                rang.interior.color = interior_color
                rang.font = font

            def report_title(font: styles.Font):
                """
                Add features to the report title cells
                :param font: 
                :return: 
                """
                rang = self.ws['{}:{}'.format(
                    '{}{}'.format(utils.get_column_letter(self.POSITION['first_sample_col']),
                                  self.POSITION['title_header_row']),
                    '{}{}'.format(utils.get_column_letter(self.POSITION['first_sample_col'] +
                                                          len(self.sample_dates) * len(self.sample_headers)),
                                  self.POSITION['title_header_row']))]
                self.ws.merge_cells(rang)
                rang.font = font

            def vehicles_model_year_merges():
                """
                Merges row of the same vehicle with different model year
                :return: 
                """
                self.ws.merge_cells(self.ws['{}:{}'.format(
                    '{}{}'.format(utils.get_column_letter(self.POSITION['vehicle_col']),
                                  self.POSITION['time_header_row']),
                    '{}{}'.format(utils.get_column_letter(self.POSITION['vehicle_col']),
                                  self.POSITION['info_header_row'])
                )])
                self.ws.merge_cells(self.ws['{}:{}'.format(
                    '{}{}'.format(utils.get_column_letter(self.POSITION['prod_model_year_col']),
                                  self.POSITION['time_header_row']),
                    '{}{}'.format(utils.get_column_letter(self.POSITION['prod_model_year_col']),
                                  self.POSITION['info_header_row']))])

            def atomic_features():
                """
                Minimal features to be done over each header
                :return: 
                """
                def merge_time_header_cells():
                    """
                    Merge time cells to make the time header appear as one for all its sub-headers
                    :return: 
                    """
                    self.ws.merge_cells(self.ws['{}:{}'.format(
                        '{}{}'.format(utils.get_column_letter(self.POSITION['first_sample_col'] +
                                                              sample_date_index * len(self.sample_headers)),
                                      self.POSITION['time_header_row']),
                        '{}{}'.format(utils.get_column_letter(self.POSITION['first_sample_col'] +
                                                              sample_date_index * len(self.sample_headers) +
                                                              max_sample_header_offset),
                                      self.POSITION['time_header_row']))])

                def paint_header_borders():
                    """
                    Paint white the border to easily spot header splits
                    :return: 
                    """
                    rang = self.ws['{}:{}'.format(
                        '{}{}'.format(utils.get_column_letter(self.POSITION['first_sample_col'] +
                                                              sample_date_index * len(self.sample_headers)),
                                      self.POSITION['time_header_row']),
                        '{}{}'.format(utils.get_column_letter(self.POSITION['first_sample_col'] +
                                                              sample_date_index * len(self.sample_headers) +
                                                              max_sample_header_offset),
                                      self.POSITION['info_header_row']))]
                    rang.border = styles.Border(left=styles.Side(color='FFFFFFFF'))  # white

                #  TODO calc max_sample_header_offset
                max_sample_header_offset = 0
                merge_time_header_cells()
                paint_header_borders()

            all_headers_font(styles.Font(name='Arial', sz=12, b=True))
            time_info_headers(border=styles.Border(top=styles.Side(color='00000000')),
                              interior_color=styles.Color('RED'),  # 147,0,4 dark red
                              font=styles.Font(color=styles.Color('WHITE')))  # 241,242,242 white
            report_title(styles.Font(sz=28))
            vehicles_model_year_merges()
            for sample_date_index in range(0, len(self.sample_dates)):
                atomic_features()

        def add_content_xl_features(xl_rows_amount: int):
            """
            Add color and font to cells with report data
            :param xl_rows_amount: 
            :return: 
            """
            for sample_date_index in range(0, len(self.sample_dates)):
                for header in self.sample_headers:
                    xl_col_letter = utils.get_column_letter(self.POSITION['first_sample_col'] + len(self.sample_headers)
                                                            * sample_date_index + header.offset)
                    self.ws['{}:{}'.format('{}{}'.format(xl_col_letter, self.POSITION['first_sample_row']),
                                           '{}{}'.format(xl_col_letter, xl_rows_amount))].number_format \
                        = header.number_format
            version_row_font = styles.Font(name='Arial', sz=12, b=True,
                                           color=styles.Color('WHITE'))  # (241,242,242) white
            version_row_interior_color = styles.Color('GRAY')
            for row in range(self.POSITION['first_sample_row'], xl_rows_amount):
                if self.matrix[row, self.vehicle_desc_mark_up_col] != 'x':
                    rang = self.ws['{}:{}'.format(
                        '{}{}'.format(utils.get_column_letter(self.POSITION['vehicle_col']), row),
                        '{}{}'.format(utils.get_column_letter(self.vehicle_desc_mark_up_col), row))]
                    rang.interior.color = version_row_interior_color
                    rang.font = version_row_font
                    #  TODO code for alternation between row colors

        def add_overall_features(xl_rows_amount: int):
            """
            Final alignments, border, split and viws freezes
            :return: 
            """
            rang = self.ws['{}:{}'.format(
                '{}{}'.format(utils.get_column_letter(self.POSITION['vehicle_col']), self.POSITION['title_header_row']),
                '{}{}'.format(utils.get_column_letter(self.POSITION['first_sample_col']), str(xl_rows_amount)))]

            rang.alignment = styles.Alignment(horizontal='center', vertical='center')
            self.ws['{}:{}'.format(
                '{}{}'.format(utils.get_column_letter(self.POSITION['vehicle_col']), self.POSITION['first_sample_row']),
                '{}{}'.format(utils.get_column_letter(self.POSITION['vehicle_col']), str(xl_rows_amount)))].alignment\
                = styles.Alignment(horizontal='left')
            self.ws.column_dimensions[utils.get_column_letter(self.vehicle_desc_mark_up_col)].hidden = True
            self.ws.freeze_panes = self.ws['{}:{}'.format(
                str(utils.get_column_letter(self.POSITION['first_sample_col'])),
                str(self.POSITION['first_sample_row']))]

        self.fill_empty_vehicle_cells(last_row_index)
        self.write_matrix_to_xl()
        add_header_xl_features()
        add_content_xl_features(last_row_index)  # TODO calc xl_rows_amount
        add_overall_features(last_row_index)
